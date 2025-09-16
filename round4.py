import flet as ft
import os
import sys
import asyncio
import time
import subprocess
import threading
import pygame
import openpyxl
from collections import Counter

_cached_workbook = None
_cached_sheets = {}



def get_column_c_values(sheet_name=None):
    """
    Reads all values from column C of the Excel workbook and returns them in a dictionary.
    Keys are row numbers, values are the cell values as strings.
    If sheet_name is provided, reads from that sheet; otherwise, uses the active sheet.
    Uses caching to avoid reloading the workbook multiple times.
    """
    global _cached_workbook, _cached_sheets
    try:
        if _cached_workbook is None:
            _cached_workbook = openpyxl.load_workbook("SF2025_PAUTAKAN_100HEARTBEAT.xlsx")
        if sheet_name:
            if sheet_name not in _cached_sheets:
                _cached_sheets[sheet_name] = _cached_workbook[sheet_name]
            sheet = _cached_sheets[sheet_name]
        else:
            sheet = _cached_workbook.active
        column_c_values = {}
        for row in range(1, sheet.max_row + 1):
            value = sheet[f"C{row}"].value
            if value is not None:
                column_c_values[row] = str(value)
        print(column_c_values)
        return column_c_values
    except Exception as e:
        print(f"Error reading column C: {e}")
        return {}

class Countdown(ft.Text):
    def __init__(self, seconds: int, heartbeat_script: str = "heartbeat.py", style=None, ref=None):
        super().__init__(value=str(seconds), style=style, ref=ref)
        self.initial_seconds = seconds
        self.seconds = seconds
        self.heartbeat_script = heartbeat_script
        self.started = False
        self.paused = False
        self.task = None

    def did_mount(self):
        # Don't start automatically - wait for manual start
        self.running = False

    def will_unmount(self):
        self.running = False
        if self.task and not self.task.done():
            self.task.cancel()

    def toggle_pause(self):
        self.paused = not self.paused

    def start(self):
        if self.task and not self.task.done():
            self.task.cancel()
        self.running = True
        self.seconds = self.initial_seconds
        self.value = str(self.seconds)
        self.update()
        self.task = self.page.run_task(self._update_timer)

    async def _update_timer(self):
        # locate your heartbeat.py next to this script
        script_path = os.path.join(os.path.dirname(__file__), self.heartbeat_script)

        try:
            while self.running and self.seconds > 0:
                await asyncio.sleep(1)
                if not self.paused:
                    # 1. Decrement and update UI
                    self.seconds -= 1
                    self.value = str(self.seconds)
                    self.update()

                    # 2. Fire off heartbeat.py in a non-blocking way
                    if os.path.isfile(script_path):
                        # Use the same Python interpreter
                        threading.Thread(target=lambda: subprocess.run([sys.executable, script_path], stdout=subprocess.PIPE, stderr=subprocess.PIPE), daemon=True).start()
                    else:
                        print(f"heartbeat script not found at: {script_path}")

            # When countdown ends
            if self.running:
                self.value = "Time's up!"
                self.update()
        finally:
            self.task = None

async def main(page: ft.Page):
    page.title = "ICI 2025 SF PAUTAKAN"

    page.window_full_screen = False
    page.window_resizable = False
    page.window_maximized = True

    # Define refs for dynamic text updates
    time_label_ref = ft.Ref[ft.Text]()
    time_value_ref = ft.Ref[ft.Text]()
    round_label_ref = ft.Ref[ft.Text]()
    round_value_ref = ft.Ref[ft.Text]()
    score_label_ref = ft.Ref[ft.Text]()
    score_value_ref = ft.Ref[ft.Text]()

    ans_value_ref = ft.Ref[ft.Text]()
    qnum_label_ref = ft.Ref[ft.Text]()
    qnum_value_ref = ft.Ref[ft.Text]()
    refdisqnumber_val_ref = ft.Ref[ft.Text]()

    # Display choices
    a1 = ft.Ref[ft.Text]()
    a2 = ft.Ref[ft.Text]()
    a3 = ft.Ref[ft.Text]()
    a4 = ft.Ref[ft.Text]()

    b1 = ft.Ref[ft.Text]()
    b2 = ft.Ref[ft.Text]()
    b3 = ft.Ref[ft.Text]()
    b4 = ft.Ref[ft.Text]()

    # Toggle states for each text
    toggled_states = {
        "a1": False,
        "a2": False,
        "a3": False,
        "a4": False,
        "b1": False,
        "b2": False,
        "b3": False,
        "b4": False,
    }

    # List for answers  
    answers = []
    display_index = 0
    ans_counter = 0
    
    # my variables
    logo_ref = ft.Ref[ft.Image]() # current logo

    circle1_ref = ft.Ref[ft.Image]() # circle logo
    circle2_ref = ft.Ref[ft.Image]() # circle logo
    circle3_ref = ft.Ref[ft.Image]() # circle logo
    circle4_ref = ft.Ref[ft.Image]() # circle logo
    circle_ic_ref = ft.Ref[ft.Image]() # circle ic logo

    correct_count = 0

    current_logo = "assets/nologo.png"
    timer_running = False
    countdown_ref = ft.Ref[Countdown]()  # Reference to countdown instance

    # score and time pointing system
    score_point_var = 2
    time_point_var = 3

    # Initialize pygame mixer for sound playback
    pygame.mixer.init()
    correct_sound = pygame.mixer.Sound("assets/sounds_correct.mp3")
    wrong_sound = pygame.mixer.Sound("assets/sounds_wrong.mp3")

    def add_score(points):
        if score_value_ref.current:
            current_score = int(score_value_ref.current.value)
            current_score += points
            score_value_ref.current.value = str(current_score)
            score_value_ref.current.update()
    def subtract_time(seconds):
        if countdown_ref.current:
            countdown_ref.current.seconds -= seconds
            countdown_ref.current.value = str(countdown_ref.current.seconds)
            countdown_ref.current.update()    

    # Cache the Excel data
    try:
        wb = openpyxl.load_workbook("SF2025_PAUTAKAN_100HEARTBEAT.xlsx")
        print(f"Sheet names: {wb.sheetnames}")
        cached_data = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            questions = {}
            answers = {}
            for row in range(1, sheet.max_row + 1):
                q_val = sheet[f"A{row}"].value
                a_val = sheet[f"B{row}"].value
                if q_val is not None:
                    questions[row] = str(q_val)
                if a_val is not None:
                    answers[row] = str(a_val)
            cached_data[sheet_name] = {'questions': questions, 'answers': answers}
        print("Workbook cached successfully")
    except Exception as e:
        print(f"Error loading workbook: {e}")
        cached_data = {}

    def clear_circles():
        nonlocal circle1_ref, circle2_ref, circle3_ref, circle4_ref
        try:
            circle1_ref.current.src = "assets/circle.png"
            circle1_ref.current.update()
            circle2_ref.current.src = "assets/circle.png"
            circle2_ref.current.update()
            circle3_ref.current.src = "assets/circle.png"
            circle3_ref.current.update()
            circle4_ref.current.src = "assets/circle.png"
            circle4_ref.current.update()
        except Exception as e:
            print(f"Error clearing circles: {e}")
    
    def count_display_circles():
        nonlocal correct_count, circle1_ref, circle2_ref, circle3_ref, circle4_ref
        try:
            if correct_count == 1:
                circle1_ref.current.src = "assets/circle_ic.png"
                circle1_ref.current.update()   
                threading.Thread(target=lambda: wrong_sound.play(), daemon=True).start() 
                subtract_time(time_point_var)         
            elif correct_count == 2:
                circle1_ref.current.src = "assets/circle_ic.png"
                circle1_ref.current.update()
                circle2_ref.current.src = "assets/circle_ic.png"
                circle2_ref.current.update()
                threading.Thread(target=lambda: wrong_sound.play(), daemon=True).start()
                subtract_time(time_point_var)
            elif correct_count == 3:
                circle1_ref.current.src = "assets/circle_ic.png"
                circle1_ref.current.update()
                circle2_ref.current.src = "assets/circle_ic.png"
                circle2_ref.current.update()
                circle3_ref.current.src = "assets/circle_ic.png"
                circle3_ref.current.update()
                threading.Thread(target=lambda: wrong_sound.play(), daemon=True).start()
                subtract_time(time_point_var)
            elif correct_count == 4:
                circle1_ref.current.src = "assets/circle_ic.png"
                circle1_ref.current.update()
                circle2_ref.current.src = "assets/circle_ic.png"
                circle2_ref.current.update()
                circle3_ref.current.src = "assets/circle_ic.png"
                circle3_ref.current.update()
                circle4_ref.current.src = "assets/circle_ic.png"
                circle4_ref.current.update()
                threading.Thread(target=lambda: correct_sound.play(), daemon=True).start()
                add_score(score_point_var)
        except Exception as e:
            print(f"Error updating circles: {e}")

    def clear_display():
        nonlocal answers, ans_counter
        a1.current.value = ""
        a1.current.update()
        a2.current.value = ""
        a2.current.update()
        a3.current.value = ""
        a3.current.update()
        a4.current.value = ""
        a4.current.update()
        b1.current.value = "A. "
        b1.current.update()
        b2.current.value = "B. "
        b2.current.update()
        b3.current.value = "C. "
        b3.current.update()
        b4.current.value = "D. "
        b4.current.update()
        ans_counter = 0
        answers = ["", "", "", "", "", ""]
        update_display()
    
    def update_display():
        print("update_display called")
        nonlocal answers, ans_counter
        sheet_name = sheet_selector(int(qnum_value_ref.current.value))
        if sheet_name in cached_data:
            b1.current.value = cached_data[sheet_name]['answers'].get(2, "")
            b1.current.update()
            b2.current.value = cached_data[sheet_name]['answers'].get(3, "")
            b2.current.update()
            b3.current.value = cached_data[sheet_name]['answers'].get(4, "")
            b3.current.update()
            b4.current.value = cached_data[sheet_name]['answers'].get(5, "")
            b4.current.update()

            a1.current.value = ""
            a1.current.update()
            a2.current.value = ""
            a2.current.update()
            a3.current.value = ""
            a3.current.update()
            a4.current.value = ""
            a4.current.update()

            ans_counter = 0
            answers = [
                cached_data[sheet_name]['answers'].get(2, ""),
                cached_data[sheet_name]['answers'].get(3, ""),
                cached_data[sheet_name]['answers'].get(4, ""),
                cached_data[sheet_name]['answers'].get(5, ""),
                cached_data[sheet_name]['answers'].get(6, ""),
                cached_data[sheet_name]['answers'].get(7, "")
            ]
            a1.current.value = cached_data[sheet_name]['questions'].get(2, "")
            a1.current.update()
            a2.current.value = cached_data[sheet_name]['questions'].get(3, "")
            a2.current.update()
            a3.current.value = cached_data[sheet_name]['questions'].get(4, "")
            a3.current.update()
            a4.current.value = cached_data[sheet_name]['questions'].get(5, "")
            a4.current.update()

        else:
            a1.current.value = ""
            a1.current.update()
            a2.current.value = ""
            a2.current.update()
            a3.current.value = ""
            a3.current.update()
            a4.current.value = ""
            a4.current.update()

            b1.current.value = ""
            b1.current.update()
            b2.current.value = ""
            b2.current.update()
            b3.current.value = ""
            b3.current.update()
            b4.current.value = ""
            b4.current.update()

            ans_counter = 0
            answers = ["", "", "", "", "", ""]

    def selector():

        nonlocal current_logo
        nonlocal cached_data
        nonlocal display_index
        nonlocal answers
        nonlocal correct_count
        print(f"Selector called with key: {key_display.value}")
        try:
            question_displayq = 0
            if refdisqnumber_val_ref.current is not None and refdisqnumber_val_ref.current.value.isdigit():
                question_displayq = int(refdisqnumber_val_ref.current.value)

            currentqnum = int(qnum_value_ref.current.value)
            if key_display.value == "Arrow Right":
                qnum_value_ref.current.value = str(currentqnum + 1)
                qnum_value_ref.current.update()
                round_value_ref.current.value = str(currentqnum + 1)
                round_value_ref.current.update()
                update_display()

            elif key_display.value == "Arrow Left":
                if currentqnum >= 1:
                    qnum_value_ref.current.value = str(currentqnum - 1)
                    qnum_value_ref.current.update()
                    round_value_ref.current.value = str(currentqnum - 1)
                    round_value_ref.current.update()
                    update_display()

            #Selecting specific cell
            if key_display.value == "Arrow Up":
                question_displayq = question_displayq + 1
                print(question_displayq)
                print("Hit: cell selector increment")
                print(refdisqnumber_val_ref.current.value)
                refdisqnumber_val_ref.current.value = str(question_displayq)
                refdisqnumber_val_ref.current.update()
            elif key_display.value == "Arrow Down":
                if question_displayq >= 1:
                    question_displayq = question_displayq - 1
                    print(question_displayq)
                    print("Hit: cell selector decrement")
                    print(refdisqnumber_val_ref.current.value)
                    refdisqnumber_val_ref.current.value = str(question_displayq)
                    refdisqnumber_val_ref.current.update()
                    #return question_displayq

           #selecting team
            elif key_display.value =="0":
                # clear the questions and answers display
                current_logo = "assets/nologo.png"
                clear_display()
                countdown_ref.update()
                countdown_ref.current.value = 100
                countdown_ref.current.toggle_pause()
                print(f"Logo set to: {current_logo}")
                update_display()
                if countdown_ref.current:
                    countdown_ref.current.start()
            elif key_display.value =="1":
                countdown_ref.current.start()                
                current_logo = "assets/fire.PNG"
                update_display()
                print(f"Logo set to: {current_logo}")

            elif key_display.value =="2":
                countdown_ref.current.start()                
                current_logo = "assets/wind.png"
                update_display()
                print(f"Logo set to: {current_logo}")
                if countdown_ref.current:
                    countdown_ref.current.start()
            elif key_display.value =="3":
                countdown_ref.current.start()
                current_logo = "assets/earth.png"
                update_display()

                print(f"Logo set to: {current_logo}")
                if countdown_ref.current:
                    countdown_ref.current.start()
            elif key_display.value =="4":
                countdown_ref.current.start()                
                current_logo = "assets/water.png"
                update_display()

                print(f"Logo set to: {current_logo}")
                if countdown_ref.current:
                    countdown_ref.current.start()
            elif key_display.value == "T":
                if countdown_ref.current:
                    countdown_ref.current.toggle_pause()
            elif key_display.value == " ": # if statement for space bar condition
                sheet_name = sheet_selector(int(qnum_value_ref.current.value))
                if sheet_name in cached_data:
                    answers = [
                        cached_data[sheet_name]['answers'].get(2, ""),
                        cached_data[sheet_name]['answers'].get(3, ""),
                        cached_data[sheet_name]['answers'].get(4, ""),
                        cached_data[sheet_name]['answers'].get(5, ""),
                        cached_data[sheet_name]['answers'].get(6, ""),
                        cached_data[sheet_name]['answers'].get(7, "")
                    ]
                # Add score_point_var to score_value_ref and time_point_var to time_value_ref
                if score_value_ref.current:
                    current_score = int(score_value_ref.current.value)
                    current_score += score_point_var
                    score_value_ref.current.value = str(current_score)
                    score_value_ref.current.update()
                    
                if countdown_ref.current:
                    countdown_ref.current.seconds += time_point_var
                    countdown_ref.current.value = str(countdown_ref.current.seconds)
                    countdown_ref.current.update()

                ans_value_ref.current.value = answers[display_index]
                ans_value_ref.current.update()
                print(f"Answer {display_index + 1} displayed: {answers[display_index]}")
                display_index = (display_index + 1) % 6


                get_column_c_values(sheet_name)
                # Play correct sound
                threading.Thread(target=lambda: correct_sound.play(), daemon=True).start()

            elif key_display.value == "Backspace":
                # Subtract time_point_var from time_value_ref
                if countdown_ref.current:
                    countdown_ref.current.seconds -= time_point_var
                    countdown_ref.current.value = str(countdown_ref.current.seconds)
                    countdown_ref.current.update()

                sheet_name = sheet_selector(int(qnum_value_ref.current.value))
                if sheet_name in cached_data:
                    row = int(refdisqnumber_val_ref.current.value)
                    ans = cached_data[sheet_name]['answers'].get(row, "No answer available")
                    ans_value_ref.current.value = ans
                    ans_value_ref.current.update()
                    print(f"Answer displayed on Backspace: {ans}")
                else:
                    ans_value_ref.current.value = "Sheet not found"
                    ans_value_ref.current.update()
                    print("Sheet not found")

                # Play wrong sound
                threading.Thread(target=lambda: wrong_sound.play(), daemon=True).start()

            elif key_display.value == "Enter":
                clear_circles()
                sheet_name = sheet_selector(int(qnum_value_ref.current.value))
                column_c_dict = get_column_c_values(sheet_name)
                if column_c_dict:
                    column_c_values = set(str(v).strip().lower() for v in column_c_dict.values() if v is not None)
                    print("Column C values (normalized):", column_c_values)
                    correct_count = 0
                    ref_map = {
                        "a1": a1,
                        "a2": a2,
                        "a3": a3,
                        "a4": a4,
                        "b1": b1,
                        "b2": b2,
                        "b3": b3,
                        "b4": b4,
                    }
                    for key, state in toggled_states.items():
                        if state:
                            text_ref = ref_map.get(key)
                            if text_ref and text_ref.current:
                                value = str(text_ref.current.value).strip().lower()
                                print(f"Toggled text {key}: '{value}'")
                                if value in column_c_values:
                                    correct_count += 1
                    ans_value_ref.current.value = f"Correct toggled: {correct_count}"
                    print(f"Correct count: {correct_count}")
                    count_display_circles()

                else:
                    ans_value_ref.current.value = "No data"
                ans_value_ref.current.update()

            print(f"Question display cue number: ", question_displayq)
        except Exception as ex:
            current = int(qnum_value_ref.current.value)
            #qnum_value_ref.current.value = "Err"

        logo_ref.current.src = current_logo

        # CALLING THE CURRENNT QUESTION NUMBER
        current_question_number()

    def sheet_selector(qnum):
        print(current_logo)

        logo_name = os.path.basename(current_logo).lower()
        sheet_mapping = {
            "fire.png": "FIRE",
            "wind.png": "WIND",
            "earth.png": "EARTH",
            "water.png": "WATER"
        }
        element = sheet_mapping.get(logo_name, "")
        current_sheet = f"R{qnum}-{element}" if element else ""
        return current_sheet


    def current_question_number():
        nonlocal cached_data
        #os._exit(0)
        sheet_name = sheet_selector(int(qnum_value_ref.current.value)) # name of the current sheet
        print(current_logo)
        print("Current sheet Name: ", sheet_name)
        if sheet_name in cached_data:
            qnum = int(qnum_value_ref.current.value)
            row = int(refdisqnumber_val_ref.current.value)
            q = cached_data[sheet_name]['questions'].get(row, None)
            # question_value_ref.current.value = f"Q{qnum}: {q}" if q else "⚠️ Cell is empty"
            print("Current sheet Name: ", sheet_name)
        else:
            #question_value_ref.current.value = f"❌ Sheet '{sheet_name}' not found"
            print("Current sheet Name: ", sheet_name)
            return




    


    def toggle_text(key):
        nonlocal toggled_states
        ref_map = {
            "a1": a1,
            "a2": a2,
            "a3": a3,
            "a4": a4,
            "b1": b1,
            "b2": b2,
            "b3": b3,
            "b4": b4,
        }
        text_ref = ref_map.get(key)
        if text_ref is None or text_ref.current is None:
            return

        current_state = toggled_states[key]

        if not current_state:
            # Toggle on: set size 40, color yellow
            text_ref.current.style = ft.TextStyle(size=40, color="yellow", weight=ft.FontWeight.BOLD)
            toggled_states[key] = True
        else:
            # Toggle off: reset size and color (default size 20, color white)
            text_ref.current.style = ft.TextStyle(size=20, color="white", weight=ft.FontWeight.BOLD)
            toggled_states[key] = False

        text_ref.current.update()
        print("Toggled states:", toggled_states)

    def on_keyboard(e: ft.KeyboardEvent):
        key_display.value = f"{e.key}"
        modifiers_display.value = f"Shift: {e.shift}, Ctrl: {e.ctrl}, Alt: {e.alt}, Meta: {e.meta}"
        print(f"Key pressed: {e.key}")
        selector()
        page.update()
    page.on_keyboard_event = on_keyboard
    key_display = ft.Text("", opacity=0.2,)

    modifiers_display = ft.Text("Modifiers: ")


    page.add(

        ft.Stack(
            expand=True,
            controls=[
                # Background image centered and scaled to fit
                ft.Image(
                    #BG_4.png for round 4
                    src="assets/BG_5.png",
                    expand=True,
                    fit=ft.ImageFit.CONTAIN,
                ),

                ft.Image(
                    src="assets/nologo.png",
                    width=40,
                    height=40,
                    fit=ft.ImageFit.CONTAIN,
                    left=1070,
                    top=120,
                    ref=logo_ref
                ),
           

                #ft.Text("Press any key...", opacity=0.5),
                key_display,
                #modifiers_display,


                # Scoreboard section
                ft.Container(
                    content=ft.Stack(
                        controls=[
                            ft.Container(
                                content=ft.Text("Question Number: ", style=ft.TextStyle(font_family="digital-7",size=20), ref=qnum_label_ref),
                                alignment=ft.alignment.top_left,
                                left=210,
                                top=140,
                            ),
                            ft.Container(
                                content=ft.Text("4", style=ft.TextStyle(font_family="digital-7",size=20), ref=qnum_value_ref),
                                alignment=ft.alignment.top_left,
                                left=370,
                                top=140,
                            ),
                            ft.Container(
                                content=ft.Text("TIME", style=ft.TextStyle(font_family="digital-7", size=30), ref=time_label_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=200,
                            ),
                            ft.Container(
                                #Countown timer
                                content=Countdown(seconds=100, heartbeat_script="heartbeat.py", style=ft.TextStyle(font_family="digital-7", size=60), ref=countdown_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=250,
                            ),
                            ft.Container(
                                content=ft.Text("ROUND", style=ft.TextStyle(font_family="digital-7", size=30), ref=round_label_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=350,
                            ),
                            ft.Container(
                                content=ft.Text("04", style=ft.TextStyle(font_family="digital-7", size=60), ref=round_value_ref),
                                alignment=ft.alignment.top_left,
                                left=80,
                                top=400,
                            ),
                            ft.Container(
                                content=ft.Text("SCORE", style=ft.TextStyle(font_family="digital-7", size=30), ref=score_label_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=500,
                            ),
                            ft.Container(
                                content=ft.Text("000", style=ft.TextStyle(font_family="digital-7", size=60), ref=score_value_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=550,
                            ),
                            ft.Container(
                                #######################
                                content=ft.Text("1", style=ft.TextStyle(font_family="digital-7", size=20), ref=refdisqnumber_val_ref),
                                alignment=ft.alignment.top_left,
                                opacity=0.0,
                                left=20,
                                top=600,
                            ),
                            ft.Container(
                                content=ft.Text("ANSWER", style=ft.TextStyle(font_family="digital-7", size=30)),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=650,
                            ),
                            ft.Container(
                                content=ft.Text("", style=ft.TextStyle(font_family="digital-7", size=60), ref=ans_value_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=700,
                            ),
                        ]
                    ),
                ),

                # Manually positioned Card panel (x = 100, y = 100)
                ft.Container(
                    content = ft.Card(
                        content=ft.Container(
                            content=ft.Stack(
                                controls=[
                                    ft.Container(
                                        content=ft.Text("ROUND 4: IDENTIFY", size=15, weight=ft.FontWeight.BOLD, text_align=ft.TextAlign.CENTER),
                                        left=0,
                                        top=10,
                                        width=930,
                                    ),
                                    ft.Container(
                                        content=ft.Text("A1 INNOVATIVE CONTROLS", size=20, weight=ft.FontWeight.BOLD, ref=a1, text_align=ft.TextAlign.CENTER),
                                        left=20,
                                        top=50,
                                        width=430,
                                        on_click=lambda e: toggle_text("a1")
                                    ),
                                    ft.Container(
                                        content=ft.Text("A2 INNOVATIVE CONTROLS", size=20, weight=ft.FontWeight.BOLD, ref=a2, text_align=ft.TextAlign.CENTER),
                                        left=20,
                                        top=126,
                                        width=430,
                                        on_click=lambda e: toggle_text("a2")
                                    ),
                                    ft.Container(
                                        content=ft.Text("A3 INNOVATIVE CONTROLS", size=20, weight=ft.FontWeight.BOLD, ref=a3, text_align=ft.TextAlign.CENTER),
                                        left=20,
                                        top=202,
                                        width=430,
                                        on_click=lambda e: toggle_text("a3")
                                    ),
                                    ft.Container(
                                        content=ft.Text("A4 INNOVATIVE CONTROLS", size=20, weight=ft.FontWeight.BOLD, ref=a4, text_align=ft.TextAlign.CENTER),
                                        left=20,
                                        top=278,
                                        width=430,
                                        on_click=lambda e: toggle_text("a4")
                                    ),

                                    #### B Column
                                    ft.Container(
                                        content=ft.Text("B1 INNOVATIVE CONTROLS", size=20, weight=ft.FontWeight.BOLD, ref=b1, text_align=ft.TextAlign.CENTER),
                                        left=450,
                                        top=50,
                                        width=430,
                                        on_click=lambda e: toggle_text("b1")
                                    ),
                                    ft.Container(
                                        content=ft.Text("B2 INNOVATIVE CONTROLS", size=20, weight=ft.FontWeight.BOLD, ref=b2, text_align=ft.TextAlign.CENTER),
                                        left=450,
                                        top=126,
                                        width=430,
                                        on_click=lambda e: toggle_text("b2")
                                    ),
                                    ft.Container(
                                        content=ft.Text("B3 INNOVATIVE CONTROLS", size=20, weight=ft.FontWeight.BOLD, ref=b3, text_align=ft.TextAlign.CENTER),
                                        left=450,
                                        top=202,
                                        width=430,
                                        on_click=lambda e: toggle_text("b3")
                                    ),
                                    ft.Container(
                                        content=ft.Text("B4 INNOVATIVE CONTROLS", size=20, weight=ft.FontWeight.BOLD, ref=b4, text_align=ft.TextAlign.CENTER),
                                        left=450,
                                        top=278,
                                        width=430,
                                        on_click=lambda e: toggle_text("b4")
                                    ),


                                    ### CIRCLE POINTS
                                    ft.Image(
                                        src="assets/circle.png",
                                        width=70,
                                        height=70,
                                        fit=ft.ImageFit.CONTAIN,
                                        left=240,
                                        top=380,
                                        ref=circle1_ref
                                    ),

                                    ft.Image(
                                    src="assets/circle.png",
                                        width=70,
                                        height=70,
                                        fit=ft.ImageFit.CONTAIN,
                                        left=390,
                                        top=380,
                                        ref=circle2_ref
                                    ),

                                    ft.Image(
                                        src="assets/circle.png",
                                        width=70,
                                        height=70,
                                        fit=ft.ImageFit.CONTAIN,
                                        left=540,
                                        top=380,
                                        ref=circle3_ref
                                    ),

                                     ft.Image(
                                        src="assets/circle.png",
                                        width=70,
                                        height=70,
                                        fit=ft.ImageFit.CONTAIN,
                                        left=690,
                                        top=380,
                                        ref=circle4_ref
                                    ),

                                ],
                                width=600,
                                height=480,
                            ),
                            width=930,
                            height=480,
                            bgcolor="#c70045",
                            border_radius=20,
                        ),
                        elevation=6
                    ),
                    left=195,  # X-axis position
                    top=160    # Y-axis position
                ),
             
               

            ]
        )
    )

ft.app(target=main)
