import flet as ft
import os
import sys
import asyncio
import time
import subprocess
import threading
import pygame
import openpyxl



class Countdown(ft.Text):
    def __init__(self, seconds: int, heartbeat_script: str = "heartbeat.py", style=None, ref=None):
        super().__init__(value=str(seconds), style=style, ref=ref)
        self.initial_seconds = seconds
        self.seconds = seconds
        self.heartbeat_script = heartbeat_script
        self.started = False
        self.paused = False
        self.task = None

    def did_mount(self):
        # Don't start automatically - wait for manual start
        self.running = False

    def will_unmount(self):
        self.running = False
        if self.task and not self.task.done():
            self.task.cancel()

    def toggle_pause(self):
        self.paused = not self.paused

    def start(self):
        if self.task and not self.task.done():
            self.task.cancel()
        self.running = True
        self.seconds = self.initial_seconds
        self.value = str(self.seconds)
        self.update()
        self.task = self.page.run_task(self._update_timer)

    async def _update_timer(self):
        # locate your heartbeat.py next to this script
        script_path = os.path.join(os.path.dirname(__file__), self.heartbeat_script)

        try:
            while self.running and self.seconds > 0:
                await asyncio.sleep(1)
                if not self.paused:
                    # 1. Decrement and update UI
                    self.seconds -= 1
                    self.value = str(self.seconds)
                    self.update()

                    # 2. Fire off heartbeat.py in a non-blocking way
                    if os.path.isfile(script_path):
                        # Use the same Python interpreter
                        threading.Thread(target=lambda: subprocess.run([sys.executable, script_path], stdout=subprocess.PIPE, stderr=subprocess.PIPE), daemon=True).start()
                    else:
                        print(f"heartbeat script not found at: {script_path}")

            # When countdown ends
            if self.running:
                self.value = "Time's up!"
                self.update()
        finally:
            self.task = None

async def main(page: ft.Page):
    page.title = "ICI 2025 SF PAUTAKAN"

    page.window_full_screen = False
    page.window_resizable = False
    page.window_maximized = True

    # Define refs for dynamic text updates
    time_label_ref = ft.Ref[ft.Text]()
    time_value_ref = ft.Ref[ft.Text]()
    round_label_ref = ft.Ref[ft.Text]()
    round_value_ref = ft.Ref[ft.Text]()
    score_label_ref = ft.Ref[ft.Text]()
    score_value_ref = ft.Ref[ft.Text]()

    ans_value_ref = ft.Ref[ft.Text]()
    qnum_label_ref = ft.Ref[ft.Text]()
    qnum_value_ref = ft.Ref[ft.Text]()
    refdisqnumber_val_ref = ft.Ref[ft.Text]()

    # Display choices
    a1 = ft.Ref[ft.Text]()
    a2 = ft.Ref[ft.Text]()
    a3 = ft.Ref[ft.Text]()
    a4 = ft.Ref[ft.Text]()
    a5 = ft.Ref[ft.Text]()
    a6 = ft.Ref[ft.Text]()

    b1 = ft.Ref[ft.Text]()
    b2 = ft.Ref[ft.Text]()
    b3 = ft.Ref[ft.Text]()
    b4 = ft.Ref[ft.Text]()
    b5 = ft.Ref[ft.Text]()
    b6 = ft.Ref[ft.Text]()

    # List for answers  
    answers = []
    display_index = 0
    ans_counter = 0
    
    # my variables
    logo_ref = ft.Ref[ft.Image]() # current logo

    current_logo = "assets/nologo.png"
    timer_running = False
    countdown_ref = ft.Ref[Countdown]()  # Reference to countdown instance

    # score and time pointing system
    score_point_var = 1
    time_point_var = 2

    # Initialize pygame mixer for sound playback
    pygame.mixer.init()
    correct_sound = pygame.mixer.Sound("assets/sounds_correct.mp3")
    wrong_sound = pygame.mixer.Sound("assets/sounds_wrong.mp3")

    def correct_answer_value():
        try:
            nonlocal ans_counter
            #
            if ans_counter == 0:
                b1.current.value = answers[0]
                b1.current.update()
            elif ans_counter == 1:
                b2.current.value = answers[1]
                b2.current.update()
            elif ans_counter == 2:
                b3.current.value = answers[2]
                b3.current.update()
            elif ans_counter == 3:
                b4.current.value = answers[3]
                b4.current.update()
            elif ans_counter == 4:
                b5.current.value = answers[4]
                b5.current.update()
            elif ans_counter == 5:
                b6.current.value = answers[5]
                b6.current.update()

            ans_counter += 1
            page.update()
        except Exception as e:
            print(f"Error in correct_answer_value: {e}")


    # Cache the Excel data
    try:
        wb = openpyxl.load_workbook("SF2025_PAUTAKAN_100HEARTBEAT.xlsx")
        print(f"Sheet names: {wb.sheetnames}")
        cached_data = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            questions = {}
            answers = {}
            for row in range(1, sheet.max_row + 1):
                q_val = sheet[f"A{row}"].value
                a_val = sheet[f"B{row}"].value
                if q_val is not None:
                    questions[row] = str(q_val)
                if a_val is not None:
                    answers[row] = str(a_val)
            cached_data[sheet_name] = {'questions': questions, 'answers': answers}
        print("Workbook cached successfully")
    except Exception as e:
        print(f"Error loading workbook: {e}")
        cached_data = {}

    def clear_display():
        nonlocal answers, ans_counter
        a1.current.value = ""
        a1.current.update()
        a2.current.value = ""
        a2.current.update()
        a3.current.value = ""
        a3.current.update()
        a4.current.value = ""
        a4.current.update()
        a5.current.value = ""
        a5.current.update()
        a6.current.value = ""
        a6.current.update()
        b1.current.value = "A. "
        b1.current.update()
        b2.current.value = "B. "
        b2.current.update()
        b3.current.value = "C. "
        b3.current.update()
        b4.current.value = "D. "
        b4.current.update()
        b5.current.value = "E. "
        b5.current.update()
        ans_counter = 0
        answers = ["", "", "", "", "", ""]
        update_display()
    
    def update_display():
        nonlocal answers, ans_counter
        sheet_name = sheet_selector(int(qnum_value_ref.current.value))
        if sheet_name in cached_data:
                b1.current.value = cached_data[sheet_name]['questions'].get(2, "")
                b1.current.update()
                b2.current.value = cached_data[sheet_name]['questions'].get(3, "")
                b2.current.update()
                b3.current.value = cached_data[sheet_name]['questions'].get(4, "")
                b3.current.update()
                b4.current.value = cached_data[sheet_name]['questions'].get(5, "")
                b4.current.update()
                b5.current.value = cached_data[sheet_name]['questions'].get(6, "")
                b5.current.update()
                b6.current.value = cached_data[sheet_name]['questions'].get(7, "")
                b6.current.update()
                a1.current.value = ""
                a1.current.update()
                a2.current.value = ""
                a2.current.update()
                a3.current.value = ""
                a3.current.update()
                a4.current.value = ""
                a4.current.update()
                a5.current.value = ""
                a5.current.update()
                a6.current.value = ""
                a6.current.update()
                ans_counter = 0
                answers = [
                    cached_data[sheet_name]['answers'].get(2, ""),
                    cached_data[sheet_name]['answers'].get(3, ""),
                    cached_data[sheet_name]['answers'].get(4, ""),
                    cached_data[sheet_name]['answers'].get(5, ""),
                    cached_data[sheet_name]['answers'].get(6, ""),
                    cached_data[sheet_name]['answers'].get(7, "")
                ]
                a1.current.value = cached_data[sheet_name]['questions'].get(2, "")
                a1.current.update()
                a2.current.value = cached_data[sheet_name]['questions'].get(3, "")
                a2.current.update()
                a3.current.value = cached_data[sheet_name]['questions'].get(4, "")
                a3.current.update()
                a4.current.value = cached_data[sheet_name]['questions'].get(5, "")
                a4.current.update()
                a5.current.value = cached_data[sheet_name]['questions'].get(6, "")
                a5.current.update()
        else:
            a1.current.value = ""
            a1.current.update()
            a2.current.value = ""
            a2.current.update()
            a3.current.value = ""
            a3.current.update()
            a4.current.value = ""
            a4.current.update()
            a5.current.value = ""
            a5.current.update()
            a6.current.value = ""
            a6.current.update()
            b1.current.value = ""
            b1.current.update()
            b2.current.value = ""
            b2.current.update()
            b3.current.value = ""
            b3.current.update()
            b4.current.value = ""
            b4.current.update()
            b5.current.value = ""
            b5.current.update()
            b6.current.value = ""
            b6.current.update()
            ans_counter = 0
            answers = ["", "", "", "", "", ""]

    def selector():

        nonlocal current_logo
        nonlocal cached_data
        nonlocal display_index
        nonlocal answers
        print(f"Selector called with key: {key_display.value}")
        try:
            question_displayq = 0
            if refdisqnumber_val_ref.current is not None and refdisqnumber_val_ref.current.value.isdigit():
                question_displayq = int(refdisqnumber_val_ref.current.value)

            currentqnum = int(qnum_value_ref.current.value)
            if key_display.value == "Arrow Right":
                qnum_value_ref.current.value = str(currentqnum + 1)
                qnum_value_ref.current.update()
                round_value_ref.current.value = str(currentqnum + 1)
                round_value_ref.current.update()
                update_display()

            elif key_display.value == "Arrow Left":
                if currentqnum >= 1:
                    qnum_value_ref.current.value = str(currentqnum - 1)
                    qnum_value_ref.current.update()
                    round_value_ref.current.value = str(currentqnum - 1)
                    round_value_ref.current.update()
                    update_display()



            #Selecting specific cell
            if key_display.value == "Arrow Up":
                question_displayq = question_displayq + 1
                print(question_displayq)
                print("Hit: cell selector increment")
                print(refdisqnumber_val_ref.current.value)
                refdisqnumber_val_ref.current.value = str(question_displayq)
                refdisqnumber_val_ref.current.update()
            elif key_display.value == "Arrow Down":
                if question_displayq >= 1:
                    question_displayq = question_displayq - 1
                    print(question_displayq)
                    print("Hit: cell selector decrement")
                    print(refdisqnumber_val_ref.current.value)
                    refdisqnumber_val_ref.current.value = str(question_displayq)
                    refdisqnumber_val_ref.current.update()
                    #return question_displayq

           #selecting team
            elif key_display.value =="0":
                # clear the questions and answers display
                current_logo = "assets/nologo.png"
                clear_display()
                countdown_ref.update()
                countdown_ref.current.value = 100
                countdown_ref.current.toggle_pause()
                print(f"Logo set to: {current_logo}")
                update_display()
                if countdown_ref.current:
                    countdown_ref.current.start()
            elif key_display.value =="1":
                countdown_ref.current.start()                
                current_logo = "assets/fire.PNG"
                update_display()
                print(f"Logo set to: {current_logo}")

            elif key_display.value =="2":
                countdown_ref.current.start()                
                current_logo = "assets/wind.png"
                update_display()
                print(f"Logo set to: {current_logo}")
                if countdown_ref.current:
                    countdown_ref.current.start()
            elif key_display.value =="3":
                countdown_ref.current.start()
                current_logo = "assets/earth.png"
                update_display()

                print(f"Logo set to: {current_logo}")
                if countdown_ref.current:
                    countdown_ref.current.start()
            elif key_display.value =="4":
                countdown_ref.current.start()                
                current_logo = "assets/water.png"
                update_display()

                print(f"Logo set to: {current_logo}")
                if countdown_ref.current:
                    countdown_ref.current.start()
            elif key_display.value == "T":
                if countdown_ref.current:
                    countdown_ref.current.toggle_pause()
            elif key_display.value == " ": # if statement for space bar condition
                sheet_name = sheet_selector(int(qnum_value_ref.current.value))
                if sheet_name in cached_data:
                    answers = [
                        cached_data[sheet_name]['answers'].get(2, ""),
                        cached_data[sheet_name]['answers'].get(3, ""),
                        cached_data[sheet_name]['answers'].get(4, ""),
                        cached_data[sheet_name]['answers'].get(5, ""),
                        cached_data[sheet_name]['answers'].get(6, ""),
                        cached_data[sheet_name]['answers'].get(7, "")
                    ]
                # Add score_point_var to score_value_ref and time_point_var to time_value_ref
                if score_value_ref.current:
                    current_score = int(score_value_ref.current.value)
                    current_score += score_point_var
                    score_value_ref.current.value = str(current_score)
                    score_value_ref.current.update()
                    correct_answer_value() # updating the answer on right  column
                if countdown_ref.current:
                    countdown_ref.current.seconds += time_point_var
                    countdown_ref.current.value = str(countdown_ref.current.seconds)
                    countdown_ref.current.update()

                ans_value_ref.current.value = answers[display_index]
                ans_value_ref.current.update()
                print(f"Answer {display_index + 1} displayed: {answers[display_index]}")
                display_index = (display_index + 1) % 6



                # Play correct sound
                threading.Thread(target=lambda: correct_sound.play(), daemon=True).start()

            elif key_display.value == "Backspace":
                # Subtract time_point_var from time_value_ref
                if countdown_ref.current:
                    countdown_ref.current.seconds -= time_point_var
                    countdown_ref.current.value = str(countdown_ref.current.seconds)
                    countdown_ref.current.update()

                sheet_name = sheet_selector(int(qnum_value_ref.current.value))
                if sheet_name in cached_data:
                    row = int(refdisqnumber_val_ref.current.value)
                    ans = cached_data[sheet_name]['answers'].get(row, "No answer available")
                    ans_value_ref.current.value = ans
                    ans_value_ref.current.update()
                    print(f"Answer displayed on Backspace: {ans}")
                else:
                    ans_value_ref.current.value = "Sheet not found"
                    ans_value_ref.current.update()
                    print("Sheet not found")

                # Play wrong sound
                threading.Thread(target=lambda: wrong_sound.play(), daemon=True).start()

            print(f"Question display cue number: ", question_displayq)
        except Exception as ex:
            current = int(qnum_value_ref.current.value)
            #qnum_value_ref.current.value = "Err"

        logo_ref.current.src = current_logo

        # CALLING THE CURRENNT QUESTION NUMBER
        current_question_number()

    def sheet_selector(qnum):
        print(current_logo)

        logo_name = os.path.basename(current_logo).lower()
        sheet_mapping = {
            "fire.png": "FIRE",
            "wind.png": "WIND",
            "earth.png": "EARTH",
            "water.png": "WATER"
        }
        element = sheet_mapping.get(logo_name, "")
        current_sheet = f"R{qnum}-{element}" if element else ""
        return current_sheet


    def current_question_number():
        nonlocal cached_data
        #os._exit(0)
        sheet_name = sheet_selector(int(qnum_value_ref.current.value)) # name of the current sheet
        print(current_logo)
        print("Current sheet Name: ", sheet_name)
        if sheet_name in cached_data:
            qnum = int(qnum_value_ref.current.value)
            row = int(refdisqnumber_val_ref.current.value)
            q = cached_data[sheet_name]['questions'].get(row, None)
            # question_value_ref.current.value = f"Q{qnum}: {q}" if q else "⚠️ Cell is empty"
            print("Current sheet Name: ", sheet_name)
        else:
            #question_value_ref.current.value = f"❌ Sheet '{sheet_name}' not found"
            print("Current sheet Name: ", sheet_name)
            return




    


    def on_keyboard(e: ft.KeyboardEvent):
        key_display.value = f"{e.key}"
        modifiers_display.value = f"Shift: {e.shift}, Ctrl: {e.ctrl}, Alt: {e.alt}, Meta: {e.meta}"
        print(f"Key pressed: {e.key}")
        selector()
        page.update()
    page.on_keyboard_event = on_keyboard
    key_display = ft.Text("", opacity=0.2,)
    
    modifiers_display = ft.Text("Modifiers: ")


    page.add(
       
        ft.Stack(
            expand=True,
            controls=[
                # Background image centered and scaled to fit
                ft.Image(
                    #BG_3.png for round 3
                    src="assets/BG_4.png",
                    expand=True,
                    fit=ft.ImageFit.CONTAIN,
                ),

                ft.Image(
                    src="assets/nologo.png",
                    width=40,
                    height=40,
                    fit=ft.ImageFit.CONTAIN,
                    left=1070,
                    top=120,
                    ref=logo_ref
                ),

                #ft.Text("Press any key...", opacity=0.5),
                key_display,
                #modifiers_display,


                # Scoreboard section
                ft.Container(
                    content=ft.Stack(
                        controls=[
                            ft.Container(
                                content=ft.Text("Question Number: ", style=ft.TextStyle(font_family="digital-7",size=20), ref=qnum_label_ref),
                                alignment=ft.alignment.top_left,
                                left=210,
                                top=140,
                            ),
                            ft.Container(
                                content=ft.Text("3", style=ft.TextStyle(font_family="digital-7",size=20), ref=qnum_value_ref),
                                alignment=ft.alignment.top_left,
                                left=370,
                                top=140,
                            ),
                            ft.Container(
                                content=ft.Text("TIME", style=ft.TextStyle(font_family="digital-7", size=30), ref=time_label_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=200,
                            ),
                            ft.Container(
                                #Countown timer
                                content=Countdown(seconds=100, heartbeat_script="heartbeat.py", style=ft.TextStyle(font_family="digital-7", size=60), ref=countdown_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=250,
                            ),
                            ft.Container(
                                content=ft.Text("ROUND", style=ft.TextStyle(font_family="digital-7", size=30), ref=round_label_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=350,
                            ),
                            ft.Container(
                                content=ft.Text("03", style=ft.TextStyle(font_family="digital-7", size=60), ref=round_value_ref),
                                alignment=ft.alignment.top_left,
                                left=80,
                                top=400,
                            ),
                            ft.Container(
                                content=ft.Text("SCORE", style=ft.TextStyle(font_family="digital-7", size=30), ref=score_label_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=500,
                            ),
                            ft.Container(
                                content=ft.Text("000", style=ft.TextStyle(font_family="digital-7", size=60), ref=score_value_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=550,
                            ),
                            ft.Container(
                                #######################
                                content=ft.Text("1", style=ft.TextStyle(font_family="digital-7", size=20), ref=refdisqnumber_val_ref),
                                alignment=ft.alignment.top_left,
                                left=20,
                                top=600,
                            ),
                            ft.Container(
                                content=ft.Text("ANSWER", style=ft.TextStyle(font_family="digital-7", size=30)),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=650,
                            ),
                            ft.Container(
                                content=ft.Text("", style=ft.TextStyle(font_family="digital-7", size=60), ref=ans_value_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=700,
                            ),
                        ]
                    ),
                ),

                # Manually positioned Card panel (x = 100, y = 100)
                ft.Container(
                    content = ft.Card(
                        content=ft.Container(
                            content=ft.Stack(
                                controls=[
                                    ft.Container(
                                        content=ft.Text("ROUND 3: PINOY UNSCRAMBLE", size=20, weight=ft.FontWeight.BOLD),
                                        left=240,
                                        top=0,
                                    ),
                                    ft.Container(
                                        content=ft.Text("A. ", size=20, weight=ft.FontWeight.BOLD, ref=a1),
                                        #alignment=ft.alignment.top_center,
                                        #opacity=0,
                                        left=230,
                                        top=40,
                                    ),
                                    ft.Container(
                                        content=ft.Text("B. ", size=20, weight=ft.FontWeight.BOLD, ref=a2),
                                        #opacity=0,
                                        left=230,
                                        top=140,
                                        #alignment=ft.alignment.top_center,
                                    ),
                                    ft.Container(
                                        content=ft.Text("C. ", size=20, weight=ft.FontWeight.BOLD, ref=a3),
                                        #opacity=0,
                                        left=230,
                                        top=240,
                                    ),
                                    ft.Container(
                                        content=ft.Text("D. ", size=20, weight=ft.FontWeight.BOLD, ref=a4),
                                        #opacity=0,
                                        left=230,
                                        top=340,
                                    ),
                                    ft.Container(
                                        content=ft.Text("E. ", size=20, weight=ft.FontWeight.BOLD, ref=a5),
                                        #opacity=0,
                                        left=230,
                                        top=440,
                                    ),


                                    #### B Column
                                    ft.Container(
                                        content=ft.Text("B1 INNOVATIVE CONTROLS", size=20, weight=ft.FontWeight.BOLD, ref=b1),
                                        #opacity=0,
                                        left=270,
                                        top=40,
                                    ),
                                    ft.Container(
                                        content=ft.Text("B2 INNOVATIVE CONTROLS", size=20, weight=ft.FontWeight.BOLD, ref=b2),
                                        left=270,
                                        top=140,
                                    ),
                                    ft.Container(
                                        content=ft.Text("B3 INNOVATIVE CONTROLS", size=20, weight=ft.FontWeight.BOLD, ref=b3),
                                        left=270,
                                        top=240,
                                    ),
                                    ft.Container(
                                        content=ft.Text("B4 INNOVATIVE CONTROLS", size=20, weight=ft.FontWeight.BOLD, ref=b4),
                                        left=270,
                                        top=340,
                                    ),
                                    ft.Container(
                                        content=ft.Text("B5 INNOVATIVE CONTROLS", size=20, weight=ft.FontWeight.BOLD, ref=b5),
                                        left=270,
                                        top=440,
                                    ),

                                ],
                                width=600,
                                height=480,
                            ),
                            width=930,
                            height=480,
                            bgcolor="#c70045",
                            border_radius=20,
                        ),
                        elevation=6
                    ),
                    left=195,  # X-axis position
                    top=160    # Y-axis position
                ),
               

            ]
        )
    )

ft.app(target=main)
