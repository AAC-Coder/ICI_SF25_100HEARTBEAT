import flet as ft
import os
import sys
import asyncio
import time
import subprocess
import threading
import pygame
import openpyxl
from collections import Counter

_cached_workbook = None
_cached_sheets = {}



class Countdown(ft.Text):
    def __init__(self, seconds: int, heartbeat_script: str = "heartbeat.py", style=None, ref=None):
        super().__init__(value=str(seconds), style=style, ref=ref)
        self.initial_seconds = seconds
        self.seconds = seconds
        self.heartbeat_script = heartbeat_script
        self.started = False
        self.paused = False
        self.task = None

    def did_mount(self):
        # Don't start automatically - wait for manual start
        self.running = False

    def will_unmount(self):
        self.running = False
        if self.task and not self.task.done():
            self.task.cancel()

    def toggle_pause(self):
        self.paused = not self.paused

    def start(self):
        if self.task and not self.task.done():
            self.task.cancel()
        self.running = True
        self.paused = False
        self.seconds = self.initial_seconds
        self.value = str(self.seconds)
        self.update()
        self.task = self.page.run_task(self._update_timer)

    async def _update_timer(self):
        # locate your heartbeat.py next to this script
        script_path = os.path.join(os.path.dirname(__file__), self.heartbeat_script)

        try:
            while self.running and self.seconds > 0:
                await asyncio.sleep(1)
                if not self.paused:
                    # 1. Decrement and update UI
                    self.seconds -= 1
                    self.value = str(self.seconds)
                    self.update()

                    # 2. Fire off heartbeat.py in a non-blocking way
                    if os.path.isfile(script_path):
                        # Use the same Python interpreter
                        threading.Thread(target=lambda: subprocess.run([sys.executable, script_path], stdout=subprocess.PIPE, stderr=subprocess.PIPE), daemon=True).start()
                    else:
                        print(f"heartbeat script not found at: {script_path}")

            # When countdown ends
            if self.running:
                self.value = "Time's up!"
                self.update()
        finally:
            self.task = None

async def main(page: ft.Page):
    page.title = "ICI 2025 SF PAUTAKAN"

    page.window_full_screen = False
    page.window_resizable = False
    page.window_maximized = True

    # Define refs for dynamic text updates
    time_label_ref = ft.Ref[ft.Text]()
    time_value_ref = ft.Ref[ft.Text]()
    round_label_ref = ft.Ref[ft.Text]()
    round_value_ref = ft.Ref[ft.Text]()
    score_label_ref = ft.Ref[ft.Text]()
    score_value_ref = ft.Ref[ft.Text]()

    ans_value_ref = ft.Ref[ft.Text]()
    qnum_label_ref = ft.Ref[ft.Text]()
    qnum_value_ref = ft.Ref[ft.Text]()
    refdisqnumber_val_ref = ft.Ref[ft.Text]()
    sheet_name_ref = ft.Ref[ft.Text]()

    # Display choices
    a1 = ft.Ref[ft.Text]()
    a2 = ft.Ref[ft.Text]()


    # Toggle states for each text
    toggled_states = {
        "a1": False,
        "a2": False,

    }

    # List for answers
    answers = []
    display_index = 0
    ans_counter = 0
    cell_counter = 1
    
    # my variables
    logo_ref = ft.Ref[ft.Image]() # current logo

    current_logo = "assets/nologo.png"
    timer_running = False
    countdown_ref = ft.Ref[Countdown]()  # Reference to countdown instance

    # score and time pointing system
    score_point_var = 3
    time_point_var = 3

    # Initialize pygame mixer for sound playback
    pygame.mixer.init()
    correct_sound = pygame.mixer.Sound("assets/sounds_correct.mp3")
    wrong_sound = pygame.mixer.Sound("assets/sounds_wrong.mp3")

    def add_score(points):
        if score_value_ref.current:
            current_score = int(score_value_ref.current.value)
            current_score += points
            score_value_ref.current.value = str(current_score)
            score_value_ref.current.update()
    def subtract_time(seconds):
        if countdown_ref.current:
            countdown_ref.current.seconds -= seconds
            countdown_ref.current.value = str(countdown_ref.current.seconds)
            countdown_ref.current.update()    

    # Cache the Excel data
    try:
        wb = openpyxl.load_workbook("SF2025_PAUTAKAN_100HEARTBEAT.xlsx")
        print(f"Sheet names: {wb.sheetnames}")
        cached_data = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            questions = {}
            answers = {}
            for row in range(1, sheet.max_row + 1):
                q_val = sheet[f"A{row}"].value
                a_val = sheet[f"B{row}"].value
                if q_val is not None:
                    questions[row] = str(q_val)
                if a_val is not None:
                    answers[row] = str(a_val)
            cached_data[sheet_name] = {'questions': questions, 'answers': answers}
        print("Workbook cached successfully")
    except Exception as e:
        print(f"Error loading workbook: {e}")
        cached_data = {}


    


    def clear_display():
        nonlocal answers, ans_counter
        a1.current.value = ""
        a1.current.update()
        a2.current.value = ""
        a2.current.update()

        ans_counter = 0
        answers = ["", "", "", "", "", ""]
        update_display()
    
    def update_display():
        print("update_display called")
        nonlocal answers, ans_counter
        sheet_name = sheet_selector()
        # Update sheet name display
        if sheet_name_ref.current:
            sheet_name_ref.current.value = sheet_name
            sheet_name_ref.current.update()
        if sheet_name in cached_data:


            a1.current.value = ""
            a1.current.update()

            ans_counter = 0
            answers = [
                cached_data[sheet_name]['answers'].get(2, ""),
                cached_data[sheet_name]['answers'].get(3, ""),
                cached_data[sheet_name]['answers'].get(4, ""),
                cached_data[sheet_name]['answers'].get(5, ""),
                cached_data[sheet_name]['answers'].get(6, ""),
                cached_data[sheet_name]['answers'].get(7, "")
            ]
            row = int(refdisqnumber_val_ref.current.value)
            a1.current.value = cached_data[sheet_name]['questions'].get(row, "")
            a1.current.update()

        else:
            a1.current.value = ""
            a1.current.update()



            ans_counter = 0
            answers = ["", "", "", "", "", ""]

    sheet_names = ["R6-CHAMPIONSHIP", "R6-BATTLE FOR 3RD"]
    current_sheet_index = 0

    def selector():
        nonlocal current_sheet_index
        nonlocal cached_data
        nonlocal display_index
        nonlocal answers
        nonlocal cell_counter
        print(f"Selector called with key: {key_display.value}")
        try:
            currentqnum = int(qnum_value_ref.current.value)
            if key_display.value == "ArrowRight":
                qnum_value_ref.current.value = str(currentqnum + 1)
                qnum_value_ref.current.update()
                round_value_ref.current.value = str(currentqnum + 1)
                round_value_ref.current.update()
                update_display()

            elif key_display.value == "ArrowLeft":
                if currentqnum >= 1:
                    qnum_value_ref.current.value = str(currentqnum - 1)
                    qnum_value_ref.current.update()
                    round_value_ref.current.value = str(currentqnum - 1)
                    round_value_ref.current.update()
                    update_display()

            # Use arrow up/down to toggle between two sheets
            if key_display.value == "Arrow Up":
                if current_sheet_index == -1:
                    current_sheet_index = 0
                else:
                    current_sheet_index = (current_sheet_index + 1) % len(sheet_names)
                
                print(f"Sheet index incremented to {current_sheet_index}")
                # Reset selected cell to 2 (A2) when sheet changes
                refdisqnumber_val_ref.current.value = "2"
                refdisqnumber_val_ref.current.update()
                update_display()
                if countdown_ref.current:
                    countdown_ref.current.start()
                a2.current.value = ""
                a2.current.update()
            elif key_display.value == "Arrow Down":
                if current_sheet_index == -1:
                    current_sheet_index = 1
                else:
                    current_sheet_index = (current_sheet_index - 1) % len(sheet_names)
                print(f"Sheet index decremented to {current_sheet_index}")
                # Reset selected cell to 2 (A2) when sheet changes
                refdisqnumber_val_ref.current.value = "2"
                refdisqnumber_val_ref.current.update()
                update_display()
                if countdown_ref.current:
                    countdown_ref.current.start()
                a2.current.value = ""
                a2.current.update()
            elif key_display.value == "0":
                # clear the questions and answers display
                clear_display()
                if countdown_ref.current:
                    countdown_ref.current.start()
            elif key_display.value == "T":
                if countdown_ref.current:
                    countdown_ref.current.toggle_pause()
            elif key_display.value == " ":  # Cycle through questions in column A from A2 to 12
                # Add score_point_var to score_value_ref
                if score_value_ref.current:
                    current_score = int(score_value_ref.current.value)
                    current_score += score_point_var
                    score_value_ref.current.value = str(current_score)
                    score_value_ref.current.update()
                # Subtract time_point_var from countdown timer
                if countdown_ref.current:
                    countdown_ref.current.seconds += time_point_var
                    countdown_ref.current.value = str(countdown_ref.current.seconds)
                    countdown_ref.current.update()
                if cell_counter < 12:
                    cell_counter += 1
                sheet_name = sheet_selector()
                if sheet_name in cached_data:
                    question = cached_data[sheet_name]['questions'].get(cell_counter, "")
                    print(f"Space bar: cell_counter = {cell_counter}, question = '{question}'")
                    if a1.current is None:
                        print("a1.current is None")
                    else:
                        a1.current.value = question
                        print(f"a1.current.value set to: '{a1.current.value}'")
                        a1.current.update()
                        page.update()  # Ensure UI refreshes after updating a1
                    refdisqnumber_val_ref.current.value = str(cell_counter)
                    refdisqnumber_val_ref.current.update()
                    answer = cached_data[sheet_name]['answers'].get(cell_counter - 1, "")
                    if a2.current is None:
                        print("a2.current is None")
                    else:
                        a2.current.value = answer
                        a2.current.update()
                        page.update()  # Ensure UI refreshes after updating a2

            elif key_display.value == "Backspace":
                # Subtract time_point_var from time_value_ref
                if countdown_ref.current:
                    countdown_ref.current.seconds -= time_point_var
                    countdown_ref.current.value = str(countdown_ref.current.seconds)
                    countdown_ref.current.update()

                sheet_name = sheet_selector()
                if sheet_name in cached_data:
                    row = int(refdisqnumber_val_ref.current.value)
                    ans = cached_data[sheet_name]['answers'].get(row, "No answer available")
                    ans_value_ref.current.value = ans
                    ans_value_ref.current.update()
                    print(f"Answer displayed on Backspace: {ans}")
                else:
                    ans_value_ref.current.value = "Sheet not found"
                    ans_value_ref.current.update()
                    print("Sheet not found")

                # Play wrong sound
                threading.Thread(target=lambda: wrong_sound.play(), daemon=True).start()

            print(f"Question display cue number: ", refdisqnumber_val_ref.current.value)
        except Exception as ex:
            current = int(qnum_value_ref.current.value)
            #qnum_value_ref.current.value = "Err"

        logo_ref.current.src = current_logo

        # CALLING THE CURRENT QUESTION NUMBER
        current_question_number()


    def sheet_selector():
        if current_sheet_index == -1:
            return ""
        return sheet_names[current_sheet_index]


    def current_question_number():
        nonlocal cached_data
        #os._exit(0)
        sheet_name = sheet_selector() # name of the current sheet
        print(current_logo)
        print("Current sheet Name: ", sheet_name)
        if sheet_name in cached_data:
            qnum = int(qnum_value_ref.current.value)
            row = int(refdisqnumber_val_ref.current.value)
            q = cached_data[sheet_name]['questions'].get(row, None)
            a1.current.value = q if q else "⚠️ Cell is empty"
            a1.current.update()
            print("Current sheet Name: ", sheet_name)
        else:
            if sheet_name == "":
                a1.current.value = ""
            else:
                a1.current.value = f"❌ Sheet '{sheet_name}' not found"
            a1.current.update()
            print("Current sheet Name: ", sheet_name)
            return



    def toggle_text(key):
        nonlocal toggled_states
        ref_map = {
            "a1": a1,

        }
        text_ref = ref_map.get(key)
        if text_ref is None or text_ref.current is None:
            return

        current_state = toggled_states[key]

        if not current_state:
            # Toggle on: set size 40, color yellow
            text_ref.current.style = ft.TextStyle(size=40, color="yellow", weight=ft.FontWeight.BOLD)
            toggled_states[key] = True
        else:
            # Toggle off: reset size and color (default size 20, color white)
            text_ref.current.style = ft.TextStyle(size=20, color="white", weight=ft.FontWeight.BOLD)
            toggled_states[key] = False

        text_ref.current.update()
        print("Toggled states:", toggled_states)

    def on_keyboard(e: ft.KeyboardEvent):
        key_display.value = f"{e.key}"
        modifiers_display.value = f"Shift: {e.shift}, Ctrl: {e.ctrl}, Alt: {e.alt}, Meta: {e.meta}"
        print(f"Key pressed: {e.key}")
        selector()
        page.update()
    page.on_keyboard_event = on_keyboard
    key_display = ft.Text("", opacity=0.2,)

    modifiers_display = ft.Text("Modifiers: ")


    page.add(

        ft.Stack(
            expand=True,
            controls=[
                # Background image centered and scaled to fit
                ft.Image(
                    #BG_4.png for round 6
                    src="assets/BG_7.png",
                    expand=True,
                    fit=ft.ImageFit.CONTAIN,
                ),

                ft.Image(
                    src="assets/nologo.png",
                    width=40,
                    height=40,
                    fit=ft.ImageFit.CONTAIN,
                    left=1070,
                    top=120,
                    ref=logo_ref
                ),


                #ft.Text("Press any key...", opacity=0.5),
                key_display,
                #modifiers_display,


                # Scoreboard section
                ft.Container(
                    content=ft.Stack(
                        controls=[
                            ft.Container(
                                content=ft.Text("Question Number: ", style=ft.TextStyle(font_family="digital-7",size=20), ref=qnum_label_ref),
                                alignment=ft.alignment.top_left,
                                left=210,
                                top=140,
                            ),
                            ft.Container(
                                ################### counter for question number
                                content=ft.Text("6", style=ft.TextStyle(font_family="digital-7",size=20), ref=qnum_value_ref),
                                alignment=ft.alignment.top_left,
                                left=370,
                                top=140,
                            ),
                            ft.Container(
                                content=ft.Text("TIME", style=ft.TextStyle(font_family="digital-7", size=30), ref=time_label_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=200,
                            ),
                            ft.Container(
                                #Countown timer
                                content=Countdown(seconds=100, heartbeat_script="heartbeat.py", style=ft.TextStyle(font_family="digital-7", size=60), ref=countdown_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=250,
                            ),
                            ft.Container(
                                content=ft.Text("ROUND", style=ft.TextStyle(font_family="digital-7", size=30), ref=round_label_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=350,
                            ),
                            ft.Container(
                                content=ft.Text("06", style=ft.TextStyle(font_family="digital-7", size=60), ref=round_value_ref),
                                alignment=ft.alignment.top_left,
                                left=80,
                                top=400,
                            ),
                            ft.Container(
                                content=ft.Text("SCORE", style=ft.TextStyle(font_family="digital-7", size=30), ref=score_label_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=500,
                            ),
                            ft.Container(
                                content=ft.Text("000", style=ft.TextStyle(font_family="digital-7", size=60), ref=score_value_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=550,
                            ),
                            ft.Container(
                                #######################
                                content=ft.Text("2", style=ft.TextStyle(font_family="digital-7", size=20), ref=refdisqnumber_val_ref),
                                alignment=ft.alignment.top_left,
                                #opacity=0.0,
                                left=20,
                                top=600,
                            ),
                            ft.Container(
                                content=ft.Text("SHEET", style=ft.TextStyle(font_family="digital-7", size=20)),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=650,
                            ),
                            ft.Container(
                                content=ft.Text("", style=ft.TextStyle(font_family="digital-7", size=30), ref=sheet_name_ref),
                                alignment=ft.alignment.top_left,
                                left=60,
                                top=680,
                            ),

                        ]
                    ),
                ),

                # Manually positioned Card panel (x = 100, y = 100)
                ft.Container(
                    content = ft.Card(
                        content=ft.Container(
                            content=ft.Stack(
                                controls=[
                                    ft.Container(
                                        content=ft.Text("ROUND 6: KEEP UP!", size=15, weight=ft.FontWeight.BOLD, text_align=ft.TextAlign.CENTER),
                                        left=0,
                                        top=10,
                                        width=930,
                                    ),


                                    ft.Container(
                                        content=ft.Text("INNOVATIVE CONTROLS SF 2025", size=80, weight=ft.FontWeight.BOLD, ref=a1, text_align=ft.TextAlign.LEFT, max_lines=None),
                                        left=20,
                                        top=80,
                                        width=890,
                                        padding=20,
                                        alignment=ft.alignment.center,
                                        on_click=lambda e: toggle_text("a1")
                                    ),
                                    ft.Container(
                                        content=ft.Text("AINNOVATIVE CONTROLS SF 2025", size=20, weight=ft.FontWeight.BOLD, ref=a2, text_align=ft.TextAlign.LEFT, max_lines=None),
                                        left=20,
                                        top=400,
                                        width=890,
                                        padding=20,
                                        alignment=ft.alignment.center,
                                        on_click=lambda e: toggle_text("a2")
                                    ),

                                ],
                                width=600,
                                height=480,
                            ),
                            width=930,
                            height=480,
                            bgcolor="#c70045",
                            border_radius=20,
                        ),
                        elevation=6
                    ),
                    left=195,  # X-axis position
                    top=160    # Y-axis position
                ),




            ]
        )
    )

    # Initial display
    #update_display()
    # No initial display - wait for arrow keys


ft.app(target=main)
